import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak, KeepTogether
from app.database import (
    get_users, get_teams, get_escalations, 
    get_announcements, get_timer_state
)
from app.db_session import get_db_status

def generate_excel_export() -> io.BytesIO:
    """
    Generates a professional multi-sheet Excel spreadsheet (.xlsx) 
    containing all Hackathon records, teams, developers, inquiries, and telemetry.
    """
    wb = Workbook()
    
    # Common Styles
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    title_font = Font(name="Arial", size=14, bold=True, color="0F172A")
    regular_font = Font(name="Arial", size=10)
    bold_font = Font(name="Arial", size=10, bold=True)
    thin_border = Border(
        left=Side(style="thin", color="E2E8F0"),
        right=Side(style="thin", color="E2E8F0"),
        top=Side(style="thin", color="E2E8F0"),
        bottom=Side(style="thin", color="E2E8F0")
    )
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")

    # ==========================================
    # SHEET 1: Overview & Telemetry
    # ==========================================
    ws_summary = wb.active
    ws_summary.title = "Event Overview"
    ws_summary.views.sheetView[0].showGridLines = True

    users = get_users()
    teams = get_teams()
    escalations = get_escalations()
    announcements = get_announcements()
    timer = get_timer_state()
    db_status = get_db_status()

    ws_summary.append(["GIETU HACKATHON 2026 - EVENT TELEMETRY & OVERVIEW"])
    ws_summary["A1"].font = title_font
    ws_summary.append(["Generated At", datetime.now().strftime("%Y-%m-%d %H:%M:%S IST")])
    ws_summary.append(["Database Engine", f"SQLite 3 Database ({db_status.get('databaseFile', 'hackathon.db')})"])
    ws_summary.append(["Sprint Duration", timer.get("durationText", "48 Hours")])
    ws_summary.append(["Timer Status", timer.get("status", "idle").upper()])
    ws_summary.append([])

    ws_summary.append(["Metric", "Count / Value", "Notes"])
    for col_idx in range(1, 4):
        cell = ws_summary.cell(row=7, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    summary_rows = [
        ("Total Registered Participants", len(users), "Includes both team members and solo hackers"),
        ("Active Teams Formed", len(teams), "Teams registered across all tracks"),
        ("Total Inquiries / Escalations", len(escalations), "Helpdesk support queries submitted by hackers"),
        ("Pending Unresolved Escalations", len([e for e in escalations if e.get("status") == "pending"]), "Awaiting admin response / resolution"),
        ("Broadcast Announcements", len(announcements), "Live notices sent to participant dashboards")
    ]

    for row_data in summary_rows:
        ws_summary.append(list(row_data))
        row_idx = ws_summary.max_row
        for col_idx in range(1, 4):
            cell = ws_summary.cell(row=row_idx, column=col_idx)
            cell.font = regular_font
            cell.border = thin_border

    # ==========================================
    # SHEET 2: Teams & Rosters
    # ==========================================
    ws_teams = wb.create_sheet(title="Teams & Rosters")
    ws_teams.views.sheetView[0].showGridLines = True
    team_headers = [
        "Team ID", "Team Name", "Track", "Invite Code", 
        "Status", "GitHub Repository URL", "Submitted At", "Member Count", "Team Leader", "Leader Email", "Team Members Details", "Description"
    ]
    ws_teams.append(team_headers)
    for col_idx in range(1, len(team_headers) + 1):
        cell = ws_teams.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    for t in teams:
        members = t.get("members", [])
        leader = next((m for m in members if m.get("isLeader")), members[0] if members else {})
        member_details = ", ".join([f"{m.get('name')} ({m.get('role')})" for m in members])
        
        row = [
            t.get("id"),
            t.get("name"),
            t.get("track"),
            t.get("inviteCode"),
            t.get("status", "not_submitted"),
            t.get("githubUrl", ""),
            t.get("submittedAt", ""),
            len(members),
            leader.get("name", "N/A"),
            leader.get("email", "N/A"),
            member_details,
            t.get("description", "")
        ]
        ws_teams.append(row)
        r_idx = ws_teams.max_row
        for col_idx in range(1, len(team_headers) + 1):
            cell = ws_teams.cell(row=r_idx, column=col_idx)
            cell.font = regular_font
            cell.border = thin_border

    # ==========================================
    # SHEET 3: Registered Developers
    # ==========================================
    ws_users = wb.create_sheet(title="Registered Developers")
    ws_users.views.sheetView[0].showGridLines = True
    user_headers = [
        "User ID", "Full Name", "Username", "Email Contact", 
        "Role Specialty", "Developer Bio", "Familiar Technologies", "In Team?", "Team ID", "Verified?"
    ]
    ws_users.append(user_headers)
    for col_idx in range(1, len(user_headers) + 1):
        cell = ws_users.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    for u in users:
        skills_str = ", ".join(u.get("skills", [])) if isinstance(u.get("skills"), list) else str(u.get("skills", ""))
        in_team_str = "YES" if u.get("teamId") else "NO (Solo)"
        verified_str = "YES" if u.get("isVerified") else "NO"
        
        row = [
            u.get("id"),
            u.get("name"),
            f"@{u.get('username')}",
            u.get("email"),
            u.get("roleTitle", "Full-Stack Developer"),
            u.get("bio", ""),
            skills_str,
            in_team_str,
            u.get("teamId") or "None",
            verified_str
        ]
        ws_users.append(row)
        r_idx = ws_users.max_row
        for col_idx in range(1, len(user_headers) + 1):
            cell = ws_users.cell(row=r_idx, column=col_idx)
            cell.font = regular_font
            cell.border = thin_border

    # ==========================================
    # SHEET 4: Support Escalations
    # ==========================================
    ws_esc = wb.create_sheet(title="Support Inquiries")
    ws_esc.views.sheetView[0].showGridLines = True
    esc_headers = [
        "Inquiry ID", "Status", "Participant Name", "Email", 
        "Team", "Question / Issue", "Proposed / Resolved Answer", "Created At", "Resolved At", "Resolved By"
    ]
    ws_esc.append(esc_headers)
    for col_idx in range(1, len(esc_headers) + 1):
        cell = ws_esc.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    for e in escalations:
        row = [
            e.get("id"),
            e.get("status", "").upper(),
            e.get("userName"),
            e.get("userEmail"),
            e.get("teamName") or "Solo",
            e.get("question"),
            e.get("proposedAnswer") or e.get("rejectionReason") or "",
            e.get("createdAt"),
            e.get("resolvedAt") or "N/A",
            e.get("resolvedBy") or "N/A"
        ]
        ws_esc.append(row)
        r_idx = ws_esc.max_row
        for col_idx in range(1, len(esc_headers) + 1):
            cell = ws_esc.cell(row=r_idx, column=col_idx)
            cell.font = regular_font
            cell.border = thin_border

    # ==========================================
    # SHEET 5: Announcements
    # ==========================================
    ws_ann = wb.create_sheet(title="Broadcast Announcements")
    ws_ann.views.sheetView[0].showGridLines = True
    ann_headers = ["ID", "Title", "Severity", "Author", "Broadcast Message", "Date & Time"]
    ws_ann.append(ann_headers)
    for col_idx in range(1, len(ann_headers) + 1):
        cell = ws_ann.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    for a in announcements:
        row = [
            a.get("id"),
            a.get("title"),
            a.get("severity", "").upper(),
            a.get("author"),
            a.get("message"),
            a.get("createdAt")
        ]
        ws_ann.append(row)
        r_idx = ws_ann.max_row
        for col_idx in range(1, len(ann_headers) + 1):
            cell = ws_ann.cell(row=r_idx, column=col_idx)
            cell.font = regular_font
            cell.border = thin_border

    # Auto-adjust column widths across all sheets
    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                val_str = str(cell.value or "")
                if len(val_str) > max_len:
                    max_len = len(val_str)
            sheet.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 45)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def generate_pdf_export() -> io.BytesIO:
    """
    Generates a PDF summary report using ReportLab.
    """
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        rightMargin=36,
        leftMargin=36,
        topMargin=36,
        bottomMargin=36
    )

    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Heading1'],
        fontSize=20,
        leading=24,
        textColor=colors.HexColor('#0f172a'),
        fontName='Helvetica-Bold',
        spaceAfter=6
    )
    
    subtitle_style = ParagraphStyle(
        'DocSubtitle',
        parent=styles['Normal'],
        fontSize=10,
        leading=14,
        textColor=colors.HexColor('#64748b'),
        spaceAfter=15
    )

    h2_style = ParagraphStyle(
        'Heading2Style',
        parent=styles['Heading2'],
        fontSize=13,
        leading=16,
        textColor=colors.HexColor('#1e293b'),
        fontName='Helvetica-Bold',
        spaceBefore=14,
        spaceAfter=8
    )

    body_style = ParagraphStyle(
        'TableBody',
        parent=styles['Normal'],
        fontSize=8.5,
        leading=11,
        textColor=colors.HexColor('#1e293b')
    )

    header_style = ParagraphStyle(
        'TableHeader',
        parent=styles['Normal'],
        fontSize=9,
        leading=12,
        fontName='Helvetica-Bold',
        textColor=colors.white
    )

    elements = []

    users = get_users()
    teams = get_teams()
    escalations = get_escalations()
    announcements = get_announcements()
    timer = get_timer_state()
    db_status = get_db_status()

    # Title & Subtitle
    elements.append(Paragraph("GIETU HACKATHON 2026 — MASTER REPORT", title_style))
    elements.append(Paragraph(
        f"Generated on {datetime.now().strftime('%B %d, %Y at %H:%M:%S IST')} | "
        f"Database: SQLite 3 ({db_status.get('databaseFile', 'hackathon.db')}) | Sprint: {timer.get('durationText', '48 Hours')}",
        subtitle_style
    ))
    elements.append(Spacer(1, 10))

    # Section 1: Executive Overview Table
    elements.append(Paragraph("1. Executive Overview & Event Telemetry", h2_style))
    overview_data = [
        [Paragraph("Metric", header_style), Paragraph("Count / Status", header_style), Paragraph("Description", header_style)],
        [Paragraph("Total Participants", body_style), Paragraph(str(len(users)), body_style), Paragraph("Registered developers & hackers", body_style)],
        [Paragraph("Formed Teams", body_style), Paragraph(str(len(teams)), body_style), Paragraph("Active rosters across all tracks", body_style)],
        [Paragraph("Support Inquiries", body_style), Paragraph(str(len(escalations)), body_style), Paragraph(f"{len([e for e in escalations if e.get('status') == 'pending'])} currently pending", body_style)],
        [Paragraph("Sprint Duration", body_style), Paragraph(timer.get("durationText", "48 Hours"), body_style), Paragraph(f"Timer status: {timer.get('status', 'idle').upper()}", body_style)],
    ]
    overview_table = Table(overview_data, colWidths=[130, 90, 320])
    overview_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e293b')),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cbd5e1')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor('#ffffff'), colors.HexColor('#f8fafc')])
    ]))
    elements.append(overview_table)
    elements.append(Spacer(1, 15))

    # Section 2: Active Teams Roster Table
    elements.append(Paragraph(f"2. Teams & Rosters ({len(teams)} Registered)", h2_style))
    if teams:
        team_table_data = [
            [Paragraph("Team Name", header_style), Paragraph("Track", header_style), Paragraph("Members", header_style), Paragraph("Status / GitHub", header_style)]
        ]
        for t in teams[:15]:
            members_str = ", ".join([m.get("name", "") for m in t.get("members", [])])
            gh_info = t.get("githubUrl", "")
            status_text = (t.get("status") or "not_submitted").upper()
            status_desc = f"<b>{status_text}</b>" + (f"<br/>{gh_info}" if gh_info else "")
            team_table_data.append([
                Paragraph(t.get("name", ""), body_style),
                Paragraph(t.get("track", ""), body_style),
                Paragraph(f"{len(t.get('members', []))} members: {members_str}", body_style),
                Paragraph(status_desc, body_style)
            ])
        team_table = Table(team_table_data, colWidths=[120, 130, 140, 150])
        team_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e293b')),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cbd5e1')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor('#ffffff'), colors.HexColor('#f8fafc')])
        ]))
        elements.append(team_table)
    else:
        elements.append(Paragraph("No teams formed yet.", body_style))
    elements.append(Spacer(1, 15))

    # Section 3: Registered Participants Table
    elements.append(Paragraph(f"3. Registered Developers ({len(users)} Hackers)", h2_style))
    user_table_data = [
        [Paragraph("Developer Name", header_style), Paragraph("Role Specialty", header_style), Paragraph("Email", header_style), Paragraph("Technologies", header_style)]
    ]
    for u in users[:25]:
        skills_str = ", ".join(u.get("skills", [])) if isinstance(u.get("skills"), list) else str(u.get("skills", ""))
        user_table_data.append([
            Paragraph(f"<b>{u.get('name')}</b> (@{u.get('username')})", body_style),
            Paragraph(u.get("roleTitle", "Full-Stack Developer"), body_style),
            Paragraph(u.get("email", ""), body_style),
            Paragraph(skills_str or "General", body_style)
        ])
    user_table = Table(user_table_data, colWidths=[140, 120, 140, 140])
    user_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e293b')),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cbd5e1')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor('#ffffff'), colors.HexColor('#f8fafc')])
    ]))
    elements.append(user_table)
    elements.append(Spacer(1, 15))

    # Section 4: Support Escalations Table
    elements.append(Paragraph(f"4. Support Inquiries & Helpdesk ({len(escalations)} Logged)", h2_style))
    if escalations:
        esc_table_data = [
            [Paragraph("Status", header_style), Paragraph("Participant", header_style), Paragraph("Inquiry Question", header_style)]
        ]
        for e in escalations[:10]:
            esc_table_data.append([
                Paragraph(f"<b>{e.get('status', '').upper()}</b>", body_style),
                Paragraph(f"{e.get('userName')}<br/>{e.get('userEmail')}", body_style),
                Paragraph(e.get("question", ""), body_style)
            ])
        esc_table = Table(esc_table_data, colWidths=[70, 130, 340])
        esc_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e293b')),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cbd5e1')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor('#ffffff'), colors.HexColor('#f8fafc')])
        ]))
        elements.append(esc_table)
    else:
        elements.append(Paragraph("No inquiries logged.", body_style))

    doc.build(elements)
    buffer.seek(0)
    return buffer
